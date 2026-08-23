"""XLSX/XLS 表格 Converter。"""

from __future__ import annotations

from pathlib import Path

from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import DocumentConversionError, MissingOptionalDependencyError
from .._stream_info import StreamInfo
from .._utils import (
    escape_heading,
    finalize_markdown,
    markdown_table,
    spreadsheet_value,
    title_from_path,
    trim_trailing_empty,
)


class SpreadsheetConverter(DocumentConverter):
    extensions = frozenset({".xlsx", ".xlsm", ".xls"})

    def convert(self, source: Path, info: StreamInfo) -> DocumentConverterResult:
        try:
            sheets = (
                _read_legacy_xls(source)
                if info.extension.casefold() == ".xls"
                else _read_openpyxl_workbook(source)
            )
        except MissingOptionalDependencyError:
            raise
        except Exception as exc:
            raise DocumentConversionError(
                f"无法读取电子表格：{source}: {exc}"
            ) from exc

        blocks: list[str] = []
        for sheet_name, rows in sheets:
            blocks.append(f"## {escape_heading(sheet_name)}")
            blocks.append(markdown_table(rows) if rows else "_空工作表_")
        content = finalize_markdown("\n\n".join(blocks), source)
        return DocumentConverterResult(
            title=title_from_path(source),
            text_content=content,
            source_path=str(source),
            converter=self.__class__.__name__,
        )


def _read_openpyxl_workbook(source: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MissingOptionalDependencyError("openpyxl", "XLSX") from exc

    workbook = load_workbook(
        source,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        return [
            (
                sheet.title,
                [
                    trim_trailing_empty(
                        [spreadsheet_value(cell.value) for cell in row]
                    )
                    for row in sheet.iter_rows()
                ],
            )
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _read_legacy_xls(source: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise MissingOptionalDependencyError("xlrd", "XLS") from exc

    workbook = xlrd.open_workbook(str(source), on_demand=True)
    try:
        return [
            (
                sheet.name,
                [
                    trim_trailing_empty(
                        [
                            spreadsheet_value(sheet.cell_value(row, col))
                            for col in range(sheet.ncols)
                        ]
                    )
                    for row in range(sheet.nrows)
                ],
            )
            for sheet in workbook.sheets()
        ]
    finally:
        workbook.release_resources()


__all__ = ["SpreadsheetConverter"]
